#!/usr/bin/env python3
"""Import a compiled-data CSV export (same 2-row header/column layout as the
codebook) into HPT Resource Index.xlsx, replacing any existing data rows.

Usage:
    python3 scripts/csv_to_xlsx.py path/to/export.csv

After importing, regenerate the site's JSON with:
    python3 scripts/xlsx_to_json.py
"""

import csv
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
XLSX_PATH = ROOT / "HPT Resource Index.xlsx"
HEADER_ROW = 2
FIRST_DATA_ROW = 3


def main():
    if len(sys.argv) != 2:
        print("Usage: python3 scripts/csv_to_xlsx.py path/to/export.csv")
        sys.exit(1)
    csv_path = Path(sys.argv[1])

    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.reader(f))
    data_rows = rows[2:]  # skip the two header rows

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Clear any existing data rows before writing the new import.
    if ws.max_row >= FIRST_DATA_ROW:
        ws.delete_rows(FIRST_DATA_ROW, ws.max_row - FIRST_DATA_ROW + 1)

    for i, row in enumerate(data_rows):
        if not row or not row[0].strip():
            continue
        for j, value in enumerate(row):
            ws.cell(row=FIRST_DATA_ROW + i, column=j + 1, value=value if value != "" else None)

    wb.save(XLSX_PATH)
    print(f"Wrote {len(data_rows)} rows from {csv_path.name} into {XLSX_PATH.name}")


if __name__ == "__main__":
    main()
